import openpyxl
from progress.bar import IncrementalBar  # progress bar
from helper_functions import DB_VKTMC
import os


def auto_add_vktmc_row():
    """
    Добавить в таблицу vktmc
    """
    wb = openpyxl.load_workbook('all_rows.xlsx')
    sheet = wb.active
    bar = IncrementalBar('Loading vktmc', max = sheet.max_row) # progress bar
    data = []
    for item in sheet:
        bar.next() # progress bar

        number_asu = int(item[1].value)
        product_name = str(item[3].value).replace('\\', '\\\\').replace("'", "\\'").strip() # Отредактированное имя продукта
        if str(item[4].value).replace('\\', '\\\\').replace("'", "\\'").strip() and item[4].value: # если True то производим поиск по данному типу если False то поиск по NULL
            product_type_name = str(item[4].value).replace('\\', '\\\\').replace("'", "\\'").strip() # Отредактированный тип продукта
            product_name_query = f"(SELECT `id` FROM `product_name` WHERE BINARY `name` = '{product_name}' AND `product_type_id` = (SELECT `id` FROM `product_type` WHERE BINARY `name` = '{product_type_name}'))"
        else:
            product_name_query = f"(SELECT `id` FROM `product_name` WHERE BINARY `name` = '{product_name}' AND `product_type_id` IS NULL)"
        ntd = f"(SELECT `id` FROM `ntd` WHERE BINARY `name` = '{item[5].value}')" if item[5].value else 'NULL' # всегда без кавычек т.к. запрос
        extra_options = str(item[6].value).strip() if item[6].value else None # могут быть кавычки т.к. varchar
        measure = f"(SELECT `id` FROM `measure` WHERE `name` = '{item[7].value}')" # всегда будет
        product_group = int(item[8].value) # в excel файле просто число
        sample_size = f"(SELECT `id` FROM `sample_size` WHERE `name` = '{item[9].value}')" if item[9].value else 'NULL' # всегда без кавычек т.к. запрос
        control_card = str(item[10].value).strip() if item[10].value else None # могут быть кавычки т.к. varchar
        article_group = f"(SELECT `id` FROM `article_group` WHERE `name` = '{str(item[11].value).strip()}')" if str(item[11].value).strip() else 'NULL' # всегда без кавычек т.к. запрос
        note = str(item[12].value).strip() if item[12].value else None # могут быть кавычки т.к. text
        item_number = int(item[13].value)# в excel файле просто число

        query = f"""
        INSERT INTO `vktmc`(`number_asu`, `product_name`, `ntd`, `extra_options`, `measure`, `product_group`, `sample_size`, `control_card`, `article_group`, `note`, `item_number`) 
        VALUES (
        {number_asu},  /* number_asu */
        {product_name_query},  /* product_name */
        {ntd},  /* ntd */
        {repr(extra_options) if extra_options else 'NULL'},  /* extra_options */
        {measure},  /* measure */
        {product_group},  /* product_group */
        {sample_size},  /* sample_size */
        {repr(control_card) if control_card else 'NULL'},  /* control_card */
        {article_group}, /* article_group */
        {repr(note) if note else 'NULL'}, /* note */
        {item_number}  /* item_number */
        )
        """
        # query_data = f"({number_asu}, {product_name_query}, {ntd}, {repr(extra_options) if extra_options else 'NULL'}, {measure}, {product_group}, {sample_size}, {repr(control_card) if control_card else 'NULL'}, {article_group}, {repr(note) if note else 'NULL'}, {item_number})"
        # data.append(query_data)

        if not DB_VKTMC.insert(query):
            print("Ошибка записи следующей строки:\n" + query + "\n(Позовите создателя кода, он поможет)")
            bar.finish() # progress bar
            break

    bar.finish() # progress bar
    
    # with open('test_query.txt', 'w') as f:
    #     f.write(f"INSERT INTO `vktmc`(`number_asu`, `product_name`, `product_type`, `ntd`, `extra_options`, `measure`, `product_group`, `sample_size`, `control_card`, `article_group`, `note`, `item_number`) VALUES "+', \n'.join(data))


def auto_add_nm_code():
    """
    Добавить в таблицу NM_code
    """
    wb = openpyxl.load_workbook('all_rows.xlsx')
    sheet = wb.active
    bar = IncrementalBar('Loading NM_code', max = sheet.max_row) # progress bar
    data = []
    for item in sheet:
        bar.next() # progress bar
        for nm_code in str(item[2].value).replace(':', ';').replace(',', ';').replace('.', ';').replace(' ', '').split(';'):
            if nm_code != 'None' and nm_code != ' ' and bool(item[2].value) and nm_code and f"({item[1].value}, {nm_code})" not in data:
                data.append(f"({item[1].value}, {nm_code})")

    DB_VKTMC.insert(f"INSERT INTO `NM_code`(`number_asu`, `nm_code`) VALUES {', '.join(data)}")
    # with open('test_query.txt', 'w') as f:
    #      f.write(f"INSERT INTO `NM_code`(`number_asu`, `nm_code`) VALUES {', '.join(data)}")
    
    bar.finish() # progress bar


def auto_add_ntd_row():
    """
    Добавить в таблицу ntd
    """
    wb = openpyxl.load_workbook('all_rows.xlsx')
    sheet = wb.active
    bar = IncrementalBar('Loading ntd', max = sheet.max_row) # progress bar
    data = set()
    for item in sheet:
        bar.next() # progress bar

        if item[5].value:
             data.add("('" + str(item[5].value).replace('\\', '\\\\').replace("'", "\\'").strip() + "')")

    DB_VKTMC.insert(f"INSERT INTO `ntd`(`name`) VALUES {', '.join(data)}")
    # with open('test_query.txt', 'w') as f:
    #     f.write(f"INSERT INTO `ntd`(`name`) VALUES {', '.join(data)}")

    bar.finish() # progress bar

def auto_add_product_name_row():
    """
    Добавить в таблицу product_name
    """
    wb = openpyxl.load_workbook('all_rows.xlsx')
    sheet = wb.active
    bar = IncrementalBar('Loading product_name', max = sheet.max_row) # progress bar
    data = set()
    for item in sheet:
        bar.next() # progress bar
        product_name = str(item[3].value).replace('\\', '\\\\').replace("'", "\\'").strip()
        if str(item[4].value).replace('\\', '\\\\').replace("'", "\\'").strip() and item[4].value:
            product_type_name = str(item[4].value).replace('\\', '\\\\').replace("'", "\\'").strip()
            data.add(f"('{product_name}', (SELECT `id` FROM `product_type` WHERE BINARY `name` = '{product_type_name}'))")
        else:
            data.add(f"('{product_name}', NULL)")

    DB_VKTMC.insert(f"INSERT INTO `product_name`(`name`, `product_type_id`) VALUES {', '.join(data)}")
    # with open('test_query.txt', 'w') as f:
    #     f.write(f"INSERT INTO `product_name`(`name`, `product_type_id`) VALUES {', '.join(data)}")

    bar.finish() # progress bar


def auto_add_product_type_row():
    """
    Добавить в таблицу product_type
    """
    wb = openpyxl.load_workbook('all_rows.xlsx')
    sheet = wb.active
    bar = IncrementalBar('Loading product_type', max = sheet.max_row) # progress bar
    data = set()
    for item in sheet:
        bar.next() # progress bar
        if str(item[4].value).replace('\\', '\\\\').replace("'", "\\'").strip() and item[4].value:
            product_type_name = str(item[4].value).replace('\\', '\\\\').replace("'", "\\'").strip()
            data.add(f"('{product_type_name}')")

    DB_VKTMC.insert(f"INSERT INTO `product_type`(`name`) VALUES {', '.join(data)}")
    # with open('test_query.txt', 'w') as f:
    #     f.write(f"INSERT INTO `product_type`(`name`) VALUES {', '.join(data)}")

    bar.finish() # progress bar


def auto_add_sample_size_row():
    """
    Добавить в таблицу sample_size
    """
    wb = openpyxl.load_workbook('all_rows.xlsx')
    sheet = wb.active
    bar = IncrementalBar('Loading sample_size', max = sheet.max_row) # progress bar
    data = set()
    for item in sheet:
        bar.next() # progress bar

        if item[9].value:
             data.add("('" + str(item[9].value).replace('\\', '\\\\').replace("'", "\\'").strip() + "')")

    DB_VKTMC.insert(f"INSERT INTO `sample_size`(`name`) VALUES {', '.join(data)}")
    # with open('test_query.txt', 'w') as f:
    #     f.write(f"INSERT INTO `ntd`(`name`) VALUES {', '.join(data)}")

    bar.finish() # progress bar


def auto_add_measure_row():
    """
    Добавить в таблицу measure
    """
    wb = openpyxl.load_workbook('all_rows.xlsx')
    sheet = wb.active
    bar = IncrementalBar('Loading measure', max = sheet.max_row) # progress bar
    data = set()
    for item in sheet:
        bar.next() # progress bar

        data.add("('" + str(item[7].value).replace('\\', '\\\\').replace("'", "\\'").strip() + "')")

    DB_VKTMC.insert(f"INSERT INTO `measure`(`name`) VALUES {', '.join(data)}")
    # with open('test_query.txt', 'w') as f:
    #     f.write(f"INSERT INTO `ntd`(`name`) VALUES {', '.join(data)}")

    bar.finish() # progress bar


def auto_add_item_number_row():
    query = f"""
    INSERT INTO `item_number` (`id`, `name`, `is_archived`) VALUES
        (1, 'Не номерные - индивидуальное штрихкодирование', 0),
        (2, 'Не номерные - тарное штрихкодирование', 0),
        (3, 'Номерные - индивидуальное штрихкодирование', 0),
        (4, 'Номерные - тарное штрихкодирование', 0);
    """
    if DB_VKTMC.insert(query):
        print("item_number complete")
    else:
        print("item_number problems")


def auto_add_product_group_row():
    query = f"""
    INSERT INTO `product_group` (`id`, `name`, `is_archived`) VALUES
        (1, 'Продукция 1 группы, подлежащей сплошному ВК (испытания 100% изделий)', 0),
        (2, 'Продукция 2 группы, подлежащей выборочному ВК (испытания определенной  выборки из партии)', 0),
        (3, 'Продукция 3 группы подлежит органолептическому контролю', 0);
    """
    if DB_VKTMC.insert(query):
        print("product_group complete")
    else:
        print("product_group problems")


def auto_add_article_group_row():
    query = f"""
        INSERT INTO `article_group` (`id`, `name`, `description`) VALUES
        (1, 'АВТО', 'Продукция для автомобильной и тракторной промышленности'),
        (2, 'АРМТРУБ', 'Арматура трубопроводная'),
        (3, 'АСБЕСТ', 'Асбестотехнические изделия'),
        (4, 'БУМПОЛИ', 'Бумажная и полиграфическая продукция'),
        (5, 'БЫТПРИБ', 'Бытовые приборы, оборудование и приспособления'),
        (6, 'БЫТХИМ', 'Бытовая химия и косметика'),
        (7, 'ВЫЧТЕХ', 'Вычислительная техника и ее комплектующие'),
        (8, 'ГОЧС', 'Продукция для обеспечения требований ГО и ЧС'),
        (9, 'ДСИЗ', 'Дерматологические средства индивидуальной защиты'),
        (10, 'ЖД', 'Продукция для оборудования метрополитена, кроме подвижного состава'),
        (11, 'ЖИВОТ', 'Продукция животного происхождения'),
        (12, 'ИНСТОСН', 'Инструменты, оснастка и материалы для них'),
        (13, 'КАНЦ', 'Канцелярская продукция'),
        (14, 'КЛЕИ', 'Клеящая продукция'),
        (15, 'КОНДЕН', 'Конденсаторы'),
        (16, 'КП', 'Кабельная продукция'),
        (17, 'КРЕПЕЖ', 'Крепеж'),
        (18, 'ЛАБ', 'Оборудование для лабораторий'),
        (19, 'ЛАКОКРАС', 'Лакокрасочная продукция'),
        (20, 'ЛЕС', 'Продукция лесопильной и деревообрабатывающей промышленности'),
        (21, 'МЕБЕЛЬ', 'Мебель'),
        (22, 'МЕД', 'Медицинская продукция'),
        (23, 'МЕТИЗЫ', 'Металлические изделия'),
        (24, 'НАСОСЫ', 'Насосное оборудование'),
        (25, 'НГП', 'Продукция нефтегазовой промышленности'),
        (26, 'НЕОРГХИМ', 'Продукция неорганической химии'),
        (27, 'ОБЩМАШ', 'Продукция общемашиностроительного применения'),
        (28, 'ОРГХИМ', 'Продукция органической химии'),
        (29, 'ПИЩ', 'Продукция пищевой промышленности'),
        (30, 'ПОДШИП', 'Подшипники'),
        (31, 'ПОДЪЕМ', 'Подъемно-погрузочная техника и ее запасные части'),
        (32, 'ПОЖ', 'Пожарное оборудование и оборудование для оповещения'),
        (33, 'ПОЛИМЕР', 'Продукция из полимеров'),
        (34, 'ПС', 'Запасные части и материалы для подвижного состава'),
        (35, 'РЕАКТИВ', 'Реактивы'),
        (36, 'РЕЗИСТОР', 'Резисторы'),
        (37, 'РЕЗТЕХ', 'Резинотехнические изделия'),
        (38, 'РМО', 'Расходные материалы для печатной техники'),
        (39, 'САНТЕХ', 'Сантехника, канализация, отопление'),
        (40, 'СВЯЗЬ', 'Продукция и оборудование для связи'),
        (41, 'СДК', 'Средства допускового контроля'),
        (42, 'СИ', 'Средства измерений'),
        (43, 'СИЗ', 'Средства индивидуальной защиты'),
        (44, 'СКОБ', 'Скобяные изделия'),
        (45, 'СМОЛЫ', 'Продукция из смол'),
        (46, 'СТЕКЛО', 'Стеклянные изделия и оптическая продукция'),
        (47, 'СТРИНС', 'Строительные инструменты'),
        (48, 'СТРМАТ', 'Строительные материалы'),
        (49, 'ТЕКСТИЛЬ', 'Продукция текстильной промышленности'),
        (50, 'ТРУБЫ', 'Трубная продукция'),
        (51, 'ФОРМ', 'Форменная одежда'),
        (52, 'ХОЗБЫТ', 'Хозяйственно-бытовые изделия'),
        (53, 'ЦВЕТМЕТ', 'Цветные металлы'),
        (54, 'ЦВЕТПРО', 'Прокат из цветных металлов'),
        (55, 'ЧЕРМЕТ', 'Черные металлы'),
        (56, 'ЧЕРПРО', 'Прокат из черных металлов'),
        (57, 'ЭЛЕКТРОН', 'Электронная продукция'),
        (58, 'ЭЛМАШ', 'Прочие электрические машины'),
        (59, 'ЭЛСВЕТ', 'Электросветовая продукция'),
        (60, 'ЭЛТЕХ', 'Электротехническая продукция');
    """
    if DB_VKTMC.insert(query):
        print("article_group complete")
    else:
        print("article_group problems")


def auto_control_card():
    """
    Добавить в таблицу control_card_files
    """
    try:
        actual_file_name = {str(_.split('_')[1]).zfill(7) : _ for _ in os.listdir(f"{os.path.abspath('../')}/control_card_files/") if 'kk' in _}
    except:
        actual_file_name = False

    if not actual_file_name:
        return 'No connecting folder!'
    
    all_control_card = {str(_[0]).zfill(7) : _[1] for _ in DB_VKTMC.select("SELECT `number_asu`, `control_card` FROM `vktmc` WHERE `control_card` IS NOT NULL")}

    for number_asu in all_control_card.keys():
        if number_asu in actual_file_name.keys():
            DB_VKTMC.update(f"UPDATE `vktmc` SET `control_card` = {repr(number_asu)} WHERE `number_asu` = {number_asu}")
            DB_VKTMC.insert(f"INSERT INTO `control_card_files`(`number_asu`, `file_name`) VALUES ({number_asu}, '{actual_file_name.get(number_asu)}')")



if __name__ == '__main__':
    try:
        choice = str(input('0 - Autoload\n1 - vktmc Row\n2 - Nm Code\n3 - Ntd Row\n4 - ProductName Row\n5 - ProductType Row\n6 - SampleSize Row\n7 - Measure Row\n8 - ItemNumber Row\n9 - ArticleGroup Row\n10 - ProductGroup Row\n11 - control card Row\ngg - Exit!\n'))
        if choice == '0':
            auto_add_item_number_row()
            auto_add_product_group_row()
            auto_add_article_group_row()
            auto_add_product_type_row()
            auto_add_product_name_row()
            auto_add_ntd_row()
            auto_add_nm_code()
            auto_add_sample_size_row()
            auto_add_measure_row()
            auto_control_card()
            auto_add_vktmc_row()
        elif choice == '1':
            auto_add_vktmc_row()
        elif choice == '2':
            auto_add_nm_code()
        elif choice == '3':
            auto_add_ntd_row()
        elif choice == '4':
            auto_add_product_name_row()
        elif choice == '5':
            auto_add_product_type_row()    
        elif choice == '6':
            auto_add_sample_size_row()  
        elif choice == '7':
            auto_add_measure_row()
        elif choice == '8':
            auto_add_item_number_row()
        elif choice == '9':
            auto_add_article_group_row()
        elif choice == '10':
            auto_add_product_group_row()
        elif choice == '11':
            auto_control_card()
        else:
            print('Ошибка')
    except TypeError as e:
        print('Ошибка!\n' + e)
    except ValueError as e:
        print('Ошибка!\n' + e)
