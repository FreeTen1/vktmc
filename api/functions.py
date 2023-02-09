from helper_functions import access_dec, add_measure_row, add_ntd_row, add_product_name_row, add_sample_size_row, DB_VKTMC, add_product_type_row, is_empty, make_row_from_data, save_edit_dec, save_report_dec, RU_ROW_NAME, QUERY_INSERT_DICT
import openpyxl
import os.path
from datetime import datetime
import pytz
from active_directory import get_user_by_login
from my_sql import my_sql
from functions_get import filter_row
import os


@access_dec
def add_vktmc_row(value_dict, login):
    """
    Добавить строку в vktmc
    """
    # Проверка на существования номера АСУ в БД
    check_number_asu = DB_VKTMC.selectOne(f"SELECT COUNT(*) FROM `vktmc` WHERE `number_asu` = {int(value_dict['number_asu'])}")
    if check_number_asu[0] != 0:
        return {'status': 400}
    
    # проверка повторов кода по нм
    mistakes = []
    for nm_code in value_dict['NM_code']:
        if DB_VKTMC.selectOne(f"SELECT COUNT(*) FROM `NM_code` WHERE `nm_code` = {nm_code} AND `number_asu` != {int(value_dict['number_asu'])}")[0]:
            mistakes.append({"nm_code": nm_code, "number_asu": [_[0] for _ in DB_VKTMC.select(f"SELECT `number_asu` FROM `NM_code` WHERE `nm_code` = {nm_code}")]})
    if mistakes:
        return {"status": 400, "mistakes": mistakes}

    # Добавление "номера АСУ" и "Код по НМ"
    if value_dict['NM_code']: # если nm_code есть, то добавляем, если нету то не добавляем
        upd_nm_code = delete_nm_code(nm_code=value_dict['NM_code'], number_asu=int(value_dict['number_asu']))
        if upd_nm_code['status'] != 200:
            return upd_nm_code

    # Добавление записи в vktmc
    number_asu = int(value_dict['number_asu'])
    
    product_name = str(value_dict['product_name']).replace('\\', '\\\\').replace("'", "\\'").strip() # Отредактированное имя продукта
    if value_dict['product_type']: # если True то производим поиск по данному типу если False то поиск по NULL
        product_type_name = str(value_dict['product_type']).replace('\\', '\\\\').replace("'", "\\'").strip() # Отредактированный тип продукта
        if is_empty(f"(SELECT `id` FROM `product_type` WHERE BINARY `name` = '{product_type_name}')"):
            add_product_type_row(product_type_name)
        product_name_query = f"(SELECT `id` FROM `product_name` WHERE BINARY `name` = '{product_name}' AND `product_type_id` = (SELECT `id` FROM `product_type` WHERE BINARY `name` = '{product_type_name}'))"
        if is_empty(product_name_query): # если связки имя - тип нету, то добавь
            add_product_name_row(product_name, product_type_name)
    else:
        product_name_query = f"(SELECT `id` FROM `product_name` WHERE BINARY `name` = '{product_name}' AND `product_type_id` IS NULL)"
        if is_empty(product_name_query): # если связки имя - тип(NULL) нету, то добавь
            add_product_name_row(product_name)
    
    if value_dict['ntd']: # если ntd передан то составь запрос на поиск его в БД, иначе присвой NULL
        ntd_query = "(SELECT `id` FROM `ntd` WHERE BINARY `name` = '{}')".format(str(value_dict['ntd']).replace('\\', '\\\\').replace("'", "\\'").strip()) # всегда без кавычек т.к. запрос
        if is_empty(ntd_query): # если такого ntd нет в базе, то добавь
            add_ntd_row(value_dict['ntd'])
    else:
        ntd_query = 'NULL'
    
    extra_options = str(value_dict['extra_options']).strip() if value_dict['extra_options'] else None # могут быть кавычки т.к. varchar
    
    measure_query = "(SELECT `id` FROM `measure` WHERE `name` = '{}')".format(value_dict['measure'].replace('\\', '\\\\').replace("'", "\\'").strip()) # всегда будет
    if is_empty(measure_query): # если единицы измерения нету в базе то добавь
        add_measure_row(value_dict['measure'])
    
    product_group = int(value_dict['product_group']) # в excel файле просто число
    
    if value_dict['sample_size']:
        sample_size_query = f"(SELECT `id` FROM `sample_size` WHERE `name` = '{value_dict['sample_size']}')" # всегда без кавычек т.к. запрос
        if is_empty(sample_size_query): # если размера выборки нету в базе то добавь
            add_sample_size_row(value_dict['sample_size'])
    else:
        sample_size_query = 'NULL'
    
    control_card = str(value_dict['control_card']).strip() if value_dict['control_card'] else None # могут быть кавычки т.к. varchar
    article_group_query = f"(SELECT `id` FROM `article_group` WHERE `name` = '{value_dict['article_group']}')" if value_dict['article_group'] else 'NULL' # isEmpty не нужен, так как это константное значение НО не факт, нужно узнать

    note = str(value_dict['note']).strip() if value_dict['note'] else None # могут быть кавычки т.к. text
    item_number = int(value_dict['item_number'])# в excel файле просто число
    if item_number not in (1, 2, 3, 4):
        return {"status": 400, "message": "Некорректный вид номенклатуры"}

    # проверка сопоставление группы продукции и карты контроля
    if product_group == 3 and bool(control_card):
        return {'status': 400, 'message': 'Ошибка соответствия группы продукции и карты контроля'}

    # Проверка соответствия вида номенклатуры и группы продукции
    if item_number == 3 and product_group != 1:
        return {'status': 400, 'message': 'Ошибка соответствия группы продукции и вида номенклатуры'}

    # Проверка соответствия группы продукции и размера выборки
    if product_group != 2 and value_dict['sample_size']:
        return {'status': 400, 'message': 'Ошибка соответствия группы продукции и размера выборки'}

    if product_group != 3 and not control_card:
        control_card = "Не требуется"

    user_id = f"(SELECT `id` FROM `users` WHERE `login` = '{login}')"
    
    query = f"""
    INSERT INTO `vktmc`(`number_asu`, `product_name`, `ntd`, `extra_options`, `measure`, `product_group`, `sample_size`, `control_card`, `article_group`, `note`, `item_number`, `user_id`) 
    VALUES (
    {number_asu},  /* number_asu */
    {product_name_query},  /* product_name */
    {ntd_query},  /* ntd */
    {repr(extra_options) if extra_options else 'NULL'},  /* extra_options */
    {measure_query},  /* measure */
    {product_group},  /* product_group */
    {sample_size_query},  /* sample_size */
    {repr(control_card) if control_card else 'NULL'},  /* control_card */
    {article_group_query}, /* article_group */
    {repr(note) if note else 'NULL'}, /* note */
    {item_number},  /* item_number */
    {user_id} /* user_id */
    )
    """
    return {'status': 200} if DB_VKTMC.insert(query) else {'status': 400}


@access_dec
@save_edit_dec
def update_vktmc_row(value_dict, login, ip):
    number_asu = int(value_dict['number_asu'])
    query_list = []
    current_row = filter_row(number_asu=value_dict["number_asu"], sort_name='number_pp', sort_by='ASC')
    if {k.lower(): [str(_) for _ in v] if type(v)==type([]) else str(v) if v else None for k, v in current_row['vktmc_row'][0].items()} == {k.lower():v for k,v in value_dict.items()}:
        return {"status": 400, "message": "Данные не изменены!"}
    if value_dict['product_name']: # если product_name есть, то добавляем вместе с типом, если нету то не добавляем
        product_name = str(value_dict['product_name']).replace('\\', '\\\\').replace("'", "\\'").strip() # Отредактированное имя продукта
        if value_dict['product_type']: # если True то производим поиск по данному типу если False то поиск по NULL
            product_type_name = str(value_dict['product_type']).replace('\\', '\\\\').replace("'", "\\'").strip() # Отредактированный тип продукта
            if is_empty(f"(SELECT `id` FROM `product_type` WHERE BINARY `name` = '{product_type_name}')"):
                add_product_type_row(product_type_name)
            product_name_query = f"(SELECT `id` FROM `product_name` WHERE BINARY `name` = '{product_name}' AND `product_type_id` = (SELECT `id` FROM `product_type` WHERE BINARY `name` = '{product_type_name}'))"
            if is_empty(product_name_query): # если связки имя - тип нету, то добавь
                add_product_name_row(product_name, product_type_name)
        else:
            product_name_query = f"(SELECT `id` FROM `product_name` WHERE BINARY `name` = '{product_name}' AND `product_type_id` IS NULL)"
            if is_empty(product_name_query): # если связки имя - тип(NULL) нету, то добавь
                add_product_name_row(product_name)
        
        query_list.append(f"`product_name`={product_name_query}")

    if value_dict['ntd']: # если ntd передан то составь запрос на поиск его в БД
        ntd_query = f"(SELECT `id` FROM `ntd` WHERE BINARY `name` = '{value_dict['ntd']}')" # всегда без кавычек т.к. запрос
        if is_empty(ntd_query): # если такого ntd нет в базе, то добавь
            add_ntd_row(value_dict['ntd'])
        query_list.append(f"`ntd`={ntd_query}")
    else:
        query_list.append(f"`ntd`=NULL")


    if value_dict['extra_options']:
        query_list.append(f"`extra_options`='{str(value_dict['extra_options']).strip()}'")
    else:
        query_list.append(f"`extra_options`=NULL")

    if value_dict['measure']:
        measure_query = f"(SELECT `id` FROM `measure` WHERE `name` = '{value_dict['measure']}')" # всегда будет
        if is_empty(measure_query):
            add_measure_row(value_dict['measure'])
        query_list.append(f"`measure`={measure_query}")
    
    if value_dict['product_group']:
        product_group = int(value_dict['product_group']) # в excel файле просто число
        query_list.append(f"`product_group`={product_group}")
    
    if value_dict['sample_size']:
        sample_size_query = f"(SELECT `id` FROM `sample_size` WHERE `name` = '{value_dict['sample_size']}')" # всегда без кавычек т.к. запрос
        if is_empty(sample_size_query):
            add_sample_size_row(value_dict['sample_size'])
        query_list.append(f"`sample_size`={sample_size_query}")
    else:
        query_list.append(f"`sample_size`=NULL")


    if value_dict.get('control_card'):
        query_list.append(f"`control_card`='{str(value_dict['control_card']).strip()}'")
    else:
        query_list.append(f"`control_card`=NULL")

    if value_dict['article_group']:
        article_group_query = f"(SELECT `id` FROM `article_group` WHERE `name` = '{value_dict['article_group']}')"
        query_list.append(f"`article_group`={article_group_query}")
    else:
        query_list.append(f"`article_group`=NULL")


    if value_dict['note']:
        query_list.append(f"`note`='{str(value_dict['note']).strip()}'")
    else:
        query_list.append(f"`note`=NULL")


    if value_dict['item_number']:
        item_number = int(value_dict['item_number']) # в excel файле просто число
        query_list.append(f"`item_number`={item_number}")


    # Проверка соответствия вида номенклатуры и группы продукции
    if item_number == 3 and product_group != 1:
        return {'status': 400, 'message': 'Ошибка соответствия вида номенклатуры и группы продукции'}
    
    # Проверка соответствия группы продукции и размера выборки
    if product_group != 2 and value_dict['sample_size']:
        return {'status': 400, 'message': 'Ошибка соответствия группы продукции и размера выборки'}
    
    main_query = f"UPDATE `vktmc` SET {', '.join(query_list)} WHERE `number_asu`={number_asu}"
    return {'status': 200} if DB_VKTMC.update(main_query) else {'status': 400}


def get_excel_asu(date):
    query = f"""
SELECT 
    vktmc.number_asu, 
    NM_code.nm_code AS 'Код по НМ', 
    product_name.name as 'Наименование изделия', 
    product_type.name as 'Тип изделия', 
    ntd.name as 'НТД', 
    extra_options, 
    article_group.name as 'Группа изделия', 
    measure.name as 'Единица измерения', 
    note, 
    vktmc.product_group AS 'Группа продукции', 
    sample_size.name AS 'Размер выборки, %', 
    control_card, 
    item_number AS 'Вид номенклатуры'

FROM `vktmc`

LEFT JOIN NM_code ON NM_code.number_asu = vktmc.number_asu /* из-за того что может отсутствовать NM_code */
JOIN product_name ON product_name.id = vktmc.product_name
LEFT JOIN product_type ON product_type.id = product_name.product_type_id /* из-за NULL в product_name.product_type_id используем LEFT JOIN */
LEFT JOIN ntd ON ntd.id = vktmc.ntd
JOIN measure ON measure.id = vktmc.measure
/* JOIN product_group ON product_group.id = vktmc.product_group /* Поменять в Группе продукции на product_group.name, тогда увидим описание группы */
LEFT JOIN sample_size ON sample_size.id = vktmc.sample_size
LEFT JOIN article_group ON article_group.id = vktmc.article_group
/* JOIN item_number ON item_number.id = vktmc.item_number /* Поменять в Вид номенклатуры на item_number.name, тогда увидим описание вида номенклатуры */
WHERE `date_time_add` LIKE '{date}%' AND vktmc.`is_archived` = 0
ORDER BY number_pp ASC
    """
    data = DB_VKTMC.select(query)
    result_data_row = {}
    for row in data: # создаю список словарей из-за возможности нескольких кодов по НМ 
        if row[0] not in result_data_row.keys():
            result_data_row[row[0]] = {
                "number_asu": str(row[0]).zfill(7),
                "nm_code": [str(row[1])] if row[1] else [],
                "product_name": row[2],
                "product_type": row[3] if row[3] else '',
                "ntd": row[4] if row[4] else '',
                "extra_options": row[5] if row[5] else '',
                "article_group": row[6] if row[6] else '',
                "measure": row[7],
                "note": row[8] if row[8] else '',
                "product_group": row[9],
                "sample_size": row[10] if row[10] else '',
                "control_card": row[11] if row[11] else '',
                "item_number": row[12]
            }
        else:
            result_data_row[row[0]]['nm_code'].append(str(row[1]))
    
    for key in result_data_row.keys():# объединяю список НМ кодов в строку НМ кодов
        result_data_row[key]['nm_code'] = ';\n'.join(result_data_row[key]['nm_code']) 

    wb = openpyxl.Workbook()
    wb.create_sheet(title = f"выгрузка за {datetime.strptime(date, '%Y-%m-%d').strftime('%d.%m.%Y')}", index = 0)
    sheet = wb[f"выгрузка за {datetime.strptime(date, '%Y-%m-%d').strftime('%d.%m.%Y')}"]
    sheet.append([
        "№ АСУ-Метро", 
        "Код по НМ", 
        "Наименование изделия", 
        "Тип изделия (серия, фирма, модель, артикул, код ОКПО и т.п.)", 
        "НТД (№ чертежа, ГОСТ, ОСТ, ТУ и т.п.)", 
        "Доп. параметры (сорт, размер, вес, рост, класс точности и др.)", 
        "Группа изделия", 
        "Единица измерения", 
        "Примечания", 
        "Группа продукции", 
        "Размер выборки, %", 
        "Карта контроля", 
        "Вид номенклатуры"])

    sheet.append([_ for _ in range(1, 14)])

    for row in list(result_data_row.values()): # ой вот тут сложно =) row - словарь, где key-имя значения, value-само значение
        sheet.append(list(row.values())) # значения словаря словаря со значениями (row)
    
    path = f"../excel_files/asu_excel_file_{date}.xlsx"
    wb.save(path)
    return os.path.exists(path)



@save_report_dec
def get_excel_file_vktmc(path_save, is_filter = False, **kwargs):
    where = ' WHERE vktmc.`is_archived` = 0 '
    if is_filter:
        replacements = {
            'number_pp': 'vktmc.number_pp', 
            'number_asu': 'vktmc.number_asu', 
            'NM_code': 'NM_code.nm_code', 
            'product_name': 'product_name.name', 
            'product_type': 'product_type.name', 
            'ntd': 'ntd.name', 
            'extra_options': 'extra_options', 
            'measure': 'measure.name', 
            'product_group': 'vktmc.product_group', 
            'sample_size': 'sample_size.name', 
            'control_card': 'control_card', 
            'article_group': 'article_group.name', 
            'note': 'note', 
            'item_number': 'item_number'
            }
        
        filter_dict = [f"{changed_name} LIKE '{kwargs[key] if key == 'article_group' else '%'+kwargs[key]+'%'}'" for key, changed_name in replacements.items() if kwargs[key] is not None]
        if not filter_dict:
            return {"status": 400, "message": "Не переданы параметры поиска"}
        else:
            where += f"AND {' AND '.join(filter_dict)}"
    
    query = f"""
SELECT 
    vktmc.number_asu, 
    NM_code.nm_code AS 'Код по НМ', 
    product_name.name as 'Наименование изделия', 
    product_type.name as 'Тип изделия', 
    ntd.name as 'НТД', 
    extra_options, 
    article_group.name as 'Группа изделия', 
    measure.name as 'Единица измерения', 
    note, 
    vktmc.product_group AS 'Группа продукции', 
    sample_size.name AS 'Размер выборки, %', 
    control_card, 
    item_number AS 'Вид номенклатуры',
    vktmc.number_pp

FROM `vktmc`

LEFT JOIN NM_code ON NM_code.number_asu = vktmc.number_asu /* из-за того что может отсутствовать NM_code */
JOIN product_name ON product_name.id = vktmc.product_name
LEFT JOIN product_type ON product_type.id = product_name.product_type_id /* из-за NULL в product_name.product_type_id используем LEFT JOIN */
LEFT JOIN ntd ON ntd.id = vktmc.ntd
JOIN measure ON measure.id = vktmc.measure
/* JOIN product_group ON product_group.id = vktmc.product_group /* Поменять в Группе продукции на product_group.name, тогда увидим описание группы */
LEFT JOIN sample_size ON sample_size.id = vktmc.sample_size
LEFT JOIN article_group ON article_group.id = vktmc.article_group
/* JOIN item_number ON item_number.id = vktmc.item_number /* Поменять в Вид номенклатуры на item_number.name, тогда увидим описание вида номенклатуры */
{where}
ORDER BY number_pp ASC
    """
    data = DB_VKTMC.select(query)
    if not data:
        return {"status": 404, "message": "Данные не найдены"}
    result_data_row = {}
    for row in data: # создаю список словарей из-за возможности нескольких кодов по НМ 
        if row[0] not in result_data_row.keys():
            result_data_row[row[0]] = {
                "number_pp": row[13],
                "number_asu": str(row[0]).zfill(7),
                "nm_code": [str(row[1])] if row[1] else [],
                "product_name": row[2],
                "product_type": row[3] if row[3] else '',
                "ntd": row[4] if row[4] else '',
                "extra_options": row[5] if row[5] else '',
                "measure": row[7],
                "product_group": row[9],
                "sample_size": row[10] if row[10] else '',
                "control_card": row[11] if row[11] else '',
                "article_group": row[6] if row[6] else '',
                "note": row[8] if row[8] else '',
                "item_number": row[12]
            }
        else:
            result_data_row[row[0]]['nm_code'].append(str(row[1]))
    
    for key in result_data_row.keys():# объединяю список НМ кодов в строку НМ кодов
        result_data_row[key]['nm_code'] = ';\n'.join(result_data_row[key]['nm_code'])

    wb = openpyxl.Workbook()
    wb.create_sheet(title = f"Выгрузка", index = 0)
    sheet = wb[f"Выгрузка"]
    if is_filter:
        head_excel_file = {k: v for k, v in kwargs.items() if kwargs.get(k) and k.lower() in RU_ROW_NAME.keys()}
        sheet.append(['Применены фильтры по столбцам'])
        for en_row_name, row_value in head_excel_file.items():
            sheet.append([f"{RU_ROW_NAME[en_row_name.lower()]} - {row_value}"])
        sheet.append([""])
    sheet.append([
        "№ п/п", 
        "№ АСУ-Метро", 
        "Код по НМ", 
        "Наименование изделия", 
        "Тип изделия (серия, фирма, модель, артикул, код ОКПО и т.п.)", 
        "НТД (№ чертежа, ГОСТ, ОСТ, ТУ и т.п.)", 
        "Доп. параметры (сорт, размер, вес, рост, класс точности и др.)", 
        "Единица измерения", 
        "Группа продукции", 
        "Размер выборки, %", 
        "Карта контроля", 
        "Группа изделия", 
        "Примечания", 
        "Вид номенклатуры"])

    sheet.append([_ for _ in range(1, 15)])

    for row in list(result_data_row.values()): # ой вот тут сложно ))) row - словарь, где key-имя значения, value-само значение
        sheet.append(list(row.values())) # значения словаря словаря со значениями (row)
    
    if not os.path.exists('../excel_files'):
        os.mkdir('../excel_files')
    wb.save(path_save)
    return {"status": 200, "path": path_save[3:]} if os.path.exists(path_save) else {"status": 400, "message": "Файл не сохранился"}


@access_dec
@save_report_dec
def get_excel_file_date(login, ip, date_from, date_to, path_save, is_archived):
    date_from = datetime.strptime(date_from, '%Y-%m-%d')
    date_to = datetime.strptime(date_to, '%Y-%m-%d')
    if date_from > date_to:
        return {"status": 400, "message": "Время начала не может быть больше чем время конца"}
    
    query = f"""
SELECT 
    vktmc.number_pp, 
    vktmc.number_asu, 
    NM_code.nm_code AS 'Код по НМ', 
    product_name.name as 'Наименование изделия', 
    product_type.name as 'Тип изделия', 
    ntd.name as 'НТД', 
    extra_options, 
    measure.name as 'Единица измерения', 
    vktmc.product_group AS 'Группа продукции', 
    sample_size.name AS 'Размер выборки, %', 
    control_card, 
    article_group.name as 'Группа изделия', 
    note, 
    item_number AS 'Вид номенклатуры'

FROM `vktmc`

LEFT JOIN NM_code ON NM_code.number_asu = vktmc.number_asu
JOIN product_name ON product_name.id = vktmc.product_name
LEFT JOIN product_type ON product_type.id = product_name.product_type_id
LEFT JOIN ntd ON ntd.id = vktmc.ntd
JOIN measure ON measure.id = vktmc.measure
LEFT JOIN sample_size ON sample_size.id = vktmc.sample_size
LEFT JOIN article_group ON article_group.id = vktmc.article_group
WHERE '{date_from}' <= {'vktmc.`date_time_delete`' if is_archived else 'vktmc.`date_time_add`'} AND {'vktmc.`date_time_delete`' if is_archived else 'vktmc.`date_time_add`'} < '{date_to}' AND vktmc.`is_archived` = {is_archived}

ORDER BY number_pp ASC
    """
    data = DB_VKTMC.select(query)
    if not data:
        return {"status": 400, "message": "Записи за выбранный промежуток дат не найдены"}
    
    wb = openpyxl.Workbook()
    wb.create_sheet(title = f"Выгрузка", index = 0)
    sheet = wb[f"Выгрузка"]
    ws = wb.active
    ws.merge_cells('A1:M1') # Объединение ячеек
    ws['A1'] = f'Позиции, добавленные в перечень ВКТМЦ с {date_from} по {date_to}' if not is_archived else f'Позиции, удалённые из перечня ВКТМЦ с {date_from} по {date_to}'

    data = make_row_from_data(data)
    for row in data:# объединяю список НМ кодов в строку НМ кодов
        row['nm_code'] = ';\n'.join([str(_) for _ in row['nm_code']]) if row['nm_code'] else None
    
    sheet.append([
        "№ п/п", 
        "№ АСУ-Метро", 
        "Код по НМ", 
        "Наименование изделия", 
        "Тип изделия (серия, фирма, модель, артикул, код ОКПО и т.п.)", 
        "НТД (№ чертежа, ГОСТ, ОСТ, ТУ и т.п.)", 
        "Доп. параметры (сорт, размер, вес, рост, класс точности и др.)", 
        "Единица измерения", 
        "Группа продукции", 
        "Размер выборки, %", 
        "Карта контроля", 
        "Группа изделия", 
        "Примечания", 
        "Вид номенклатуры"])

    for row in data:
        sheet.append(list(row.values()))

    if not os.path.exists('../excel_files'):
        os.mkdir('../excel_files')
    wb.save(path_save)
    return {"status": 200, "path": path_save[3:]} if os.path.exists(path_save) else {"status": 400, "message": "Файл не сохранился"}


@access_dec
@save_report_dec
def get_excel_file_date_change_row(login, ip, date_from, date_to, path_save):
    date_from = datetime.strptime(date_from, '%Y-%m-%d')
    date_to = datetime.strptime(date_to, '%Y-%m-%d')
    if date_from > date_to:
        return {"status": 400, "message": "Время начала не может быть больше чем время конца"}
    
    query = f"""
SELECT 
    vktmc.number_pp, 
    vktmc.number_asu, 
    NM_code.nm_code AS 'Код по НМ', 
    product_name.name as 'Наименование изделия', 
    product_type.name as 'Тип изделия', 
    ntd.name as 'НТД', 
    extra_options, 
    measure.name as 'Единица измерения', 
    IF (logs_change_rows.row_name = 'product_group', logs_change_rows.old_value, NULL) AS 'Группа продукции было', 
    IF (logs_change_rows.row_name = 'product_group', logs_change_rows.new_value, NULL) AS 'Группа продукции стало',
    IF (logs_change_rows.row_name = 'sample_size', logs_change_rows.old_value, NULL) AS 'Размер выборки, %, было',
    IF (logs_change_rows.row_name = 'sample_size', logs_change_rows.new_value, NULL) AS 'Размер выборки, %, стало',
    control_card, 
    article_group.name as 'Группа изделия', 
    note, 
    item_number AS 'Вид номенклатуры'

FROM `vktmc`

LEFT JOIN NM_code ON NM_code.number_asu = vktmc.number_asu
JOIN logs_change_rows ON logs_change_rows.number_pp = vktmc.number_pp AND logs_change_rows.row_name in ('product_group', 'sample_size')
JOIN product_name ON product_name.id = vktmc.product_name
LEFT JOIN product_type ON product_type.id = product_name.product_type_id
LEFT JOIN ntd ON ntd.id = vktmc.ntd
JOIN measure ON measure.id = vktmc.measure
LEFT JOIN sample_size ON sample_size.id = vktmc.sample_size
LEFT JOIN article_group ON article_group.id = vktmc.article_group
WHERE '{date_from}' <= logs_change_rows.datetime_edit AND logs_change_rows.datetime_edit < '{date_to}' AND vktmc.`is_archived` = 0

ORDER BY number_pp ASC
    """
    data = DB_VKTMC.select(query)
    if not data:
        return {"status": 400, "message": "Записи за выбранный промежуток дат не найдены"}

    result = {}
    for row in data:
        if f"{row[0]} {row[8]} {row[9]} {row[10]} {row[11]}" not in result.keys():
            result[f"{row[0]} {row[8]} {row[9]} {row[10]} {row[11]}"] = {
                "number_pp": row[0],
                "number_asu": str(row[1]).zfill(7),
                "nm_code": [row[2]] if row[2] else None,
                "product_name": row[3],
                "product_type": row[4] if row[4] else None,
                "ntd": row[5] if row[5] else None,
                "extra_options": row[6] if row[6] else None,
                "measure": row[7],
                "product_group_old": row[8] if row[8] else None,
                "product_group_new": row[9] if row[9] else None,
                "sample_size_old": row[10] if row[10] else None,
                "sample_size_new": row[11] if row[11] else None,
                "control_card": row[12] if row[12] else None,
                "article_group": row[13] if row[13] else None,
                "note": row[14] if row[14] else None,
                "item_number": row[15]
            }
        else:
            result[f"{row[0]} {row[8]} {row[9]} {row[10]} {row[11]}"]['nm_code'].append(row[2])
            
    result = list(result.values())

    for row in result:# объединяю список НМ кодов в строку НМ кодов
        row['nm_code'] = ';\n'.join([str(_) for _ in row['nm_code']]) if row['nm_code'] else None
    

    wb = openpyxl.Workbook()
    wb.create_sheet(title = f"Выгрузка", index = 0)
    sheet = wb[f"Выгрузка"]
    ws = wb.active
    ws.merge_cells('A1:M1') # Объединение ячеек
    ws['A1'] = f'Изменение группы продукции и/или выборки в перечне ВКТМЦ с {date_from} по {date_to}'
    sheet.append([
        "№ п/п", 
        "№ АСУ-Метро", 
        "Код по НМ", 
        "Наименование изделия", 
        "Тип изделия (серия, фирма, модель, артикул, код ОКПО и т.п.)", 
        "НТД (№ чертежа, ГОСТ, ОСТ, ТУ и т.п.)", 
        "Доп. параметры (сорт, размер, вес, рост, класс точности и др.)", 
        "Единица измерения", 
        "Группа продукции было", 
        "Группа продукции стало", 
        "Размер выборки, %, было", 
        "Размер выборки, %, стало", 
        "Карта контроля", 
        "Группа изделия", 
        "Примечания", 
        "Вид номенклатуры"])

    for row in result:
        sheet.append(list(row.values()))

    if not os.path.exists('../excel_files'):
        os.mkdir('../excel_files')
    wb.save(path_save)
    return {"status": 200, "path": path_save[3:]} if os.path.exists(path_save) else {"status": 400, "message": "Файл не сохранился"}


def add_new_manual(table_name, name, description):
    check = QUERY_INSERT_DICT.get(f"{table_name}_check")
    if check:
        if DB_VKTMC.selectOne(check.format(name))[0] != 0:
            return False
    query = QUERY_INSERT_DICT.get(table_name, None).format(name, description)
    return DB_VKTMC.insert(query)


def delete_and_come_back_manual(query):
    return DB_VKTMC.update(query)


def delete_nm_code(**args):
    if not args['nm_code']:
        DB_VKTMC.update(f"DELETE FROM `NM_code` WHERE `number_asu` = {int(args['number_asu'])}")
        return {"status": 200}
    
    mistakes = []
    for nm_code in args['nm_code']:
        if DB_VKTMC.selectOne(f"SELECT COUNT(*) FROM `NM_code` WHERE `nm_code` = {nm_code} AND `number_asu` != {int(args['number_asu'])}")[0]:
            mistakes.append({"nm_code": nm_code, "number_asu": [str(_[0]).zfill(7) for _ in DB_VKTMC.select(f"SELECT `number_asu` FROM `NM_code` WHERE `nm_code` = {nm_code}")]})
    
    if mistakes:
        return {"status": 400, "mistakes": mistakes}
    else:
        DB_VKTMC.update(f"DELETE FROM `NM_code` WHERE `number_asu` = {int(args['number_asu'])}")
        if args['nm_code']:
            NM_code_list = [f"({int(args['number_asu'])}, {_})" for _ in args['nm_code']]
            DB_VKTMC.insert(f"INSERT INTO `NM_code`(`number_asu`, `nm_code`) VALUES {', '.join(NM_code_list)}")
        return {"status": 200}


def authorization(login):
    db_vktmc_auth = my_sql()
    body = get_user_by_login(login) # Обращение в АД (только в сети метрополитена) РАСКОМЕНТИРОВАТЬ
    # body = {"ad_user": login, "email": "example@mosmetro.ru", "fio": "Заглушка поменять в ф-ции authorization()"} # Заглушка (закомментировать)
    if not body:
        return {"status": 404, "message": "Пользователь не найден!"}
    user_id = db_vktmc_auth.selectOne(f"SELECT `id` FROM `users` WHERE `login` = '{login}'")
    if not user_id:
        db_vktmc_auth.insert(f"INSERT INTO `users`(`login`, `email`, `full_name`) VALUES ('{body['ad_user']}', '{body['email']}', '{body['fio']}')") # создание пользователя
        user_id = db_vktmc_auth.lastid[0]
        db_vktmc_auth.insert(f"INSERT INTO `settings`(`user_id`) VALUES ({user_id})") # создание дефолтных настроек пользователя
        data = db_vktmc_auth.selectOne(f"SELECT access, settings.number_display_row FROM users JOIN settings ON settings.user_id = users.id WHERE users.id = {user_id}")
        body.update(status=200, access=data[0], number_display_row=data[1])
    else:
        user_id = user_id[0]
        db_vktmc_auth.update(f"UPDATE `users` SET `last_login_date`='{datetime.now(pytz.timezone('Europe/Moscow')).strftime('%Y-%m-%d %H:%M:%S')}' WHERE `id` = {user_id}")
        data = db_vktmc_auth.selectOne(f"SELECT access, settings.number_display_row FROM users JOIN settings ON settings.user_id = users.id WHERE users.id = {user_id}")
        body.update(status=200, access=data[0], number_display_row=data[1])
    
    return body


def save_settings(**kwargs):
    query = [f"`{key}` = '{value}'" for key, value in kwargs.items() if key != 'login']
    return {"status":200} if DB_VKTMC.update(f"UPDATE `settings` SET {', '.join(query)} WHERE `user_id` = (SELECT `id` FROM users WHERE login = '{kwargs['login']}')") else {"message": "Ошибка сохранения настройки 'кол-во выводимых строк'", "status":400}


@access_dec
def toggle_delete_vktmc_row(number_pp, login):
    check = DB_VKTMC.selectOne(f"SELECT `is_archived` FROM `vktmc` WHERE `number_pp` = {number_pp}")[0] # проверка
    if check:
        if DB_VKTMC.update(f"UPDATE `vktmc` SET `is_archived`= 0, `date_time_add` = '{datetime.now(pytz.timezone('Europe/Moscow')).strftime('%Y-%m-%d %H:%M:%S')}', `user_id` = (SELECT `id` FROM `users` WHERE `login` = '{login}') WHERE `number_pp` = {number_pp}"): # вернуть
            return {"status": 200, "message": 'Строка возвращена из архива!'}
        else:
            {"status": 400, "message": 'Ошибка возвращения!'}
    else:
        if DB_VKTMC.update(f"UPDATE `vktmc` SET `is_archived`= 1, `date_time_delete` = '{datetime.now(pytz.timezone('Europe/Moscow')).strftime('%Y-%m-%d %H:%M:%S')}', `user_delete_id` = (SELECT `id` FROM `users` WHERE `login` = '{login}') WHERE `number_pp` = {number_pp}") and DB_VKTMC.update(f"DELETE FROM `NM_code` WHERE `number_asu` = (SELECT vktmc.number_asu FROM vktmc WHERE vktmc.number_pp = {number_pp})"): # удалить
            return {"status": 200, "message": 'Строка удалена!'}
        else:
            {"status": 400, "message": 'Ошибка удаления!'}


@access_dec
def update_control_card(login):
    
    DB_VKTMC.update("UPDATE `vktmc` SET `control_card` = 'нет' WHERE `product_group` != 3 AND `control_card` IS NULL AND `article_group` != (SELECT `id` FROM `article_group` WHERE `name` = 'СИЗ')")
    DB_VKTMC.update("UPDATE `vktmc` SET `control_card` = 'Не требуется' WHERE `product_group` != 3 AND `control_card` IS NULL AND `article_group` = (SELECT `id` FROM `article_group` WHERE `name` = 'СИЗ')")
    
    DB_VKTMC.update("UPDATE `vktmc` SET `control_card` = NULL WHERE `product_group` = 3")
    
    # Создание словаря {number_asu: file_name}
    try:
        actual_file_name = {str(_.split('_')[1]).zfill(7) : _ for _ in os.listdir(f"{os.path.abspath('../')}/control_card_files/") if 'kk' in _}
    except:
        actual_file_name = False

    if not actual_file_name:
        return {"message": "Нет соединения с общей папкой!", "status": 400}
    
    curr_file_name = {str(_[0]).zfill(7) : _[1] for _ in DB_VKTMC.select("SELECT `number_asu`, `file_name` FROM `control_card_files`")}
    all_control_card = {str(_[0]).zfill(7) : _[1] for _ in DB_VKTMC.select("SELECT `number_asu`, `control_card` FROM `vktmc` WHERE `control_card` IS NOT NULL")}

    for number_asu, curr_control_card in all_control_card.items():
        try:
            if number_asu in actual_file_name.keys() and curr_control_card.upper() in ('НЕ ТРЕБУЕТСЯ', 'НЕТ'):
                DB_VKTMC.update(f"UPDATE `vktmc` SET `control_card` = {repr(number_asu)} WHERE `number_asu` = {number_asu}")
                DB_VKTMC.insert(f"INSERT INTO `control_card_files`(`number_asu`, `file_name`) VALUES ({number_asu}, '{actual_file_name.get(number_asu)}')")
            elif curr_control_card.upper() not in ('НЕ ТРЕБУЕТСЯ', 'НЕТ') and curr_file_name[number_asu] != actual_file_name[number_asu]:
                DB_VKTMC.update(f"UPDATE `control_card_files` SET `file_name` = {repr(actual_file_name.get(number_asu))} WHERE `number_asu` = {number_asu}")
        except KeyError as e:
            return {"status": 400, "message": f"Ошибка в позиции с номером АСУ: '{number_asu}'"}

    return {"status": 200}